from __future__ import annotations

import json
from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import Database
from app.database import ReviewDatabase


def export_findings(db: Database, project_id: int) -> bytes:
    projects = db.query("SELECT * FROM projects WHERE id=?", (project_id,))
    if not projects: raise ValueError("项目不存在")
    project = projects[0]
    findings = db.query("SELECT * FROM findings WHERE project_id=? ORDER BY id", (project_id,))
    documents = db.query("SELECT * FROM documents WHERE project_id=?", (project_id,))
    standards = db.query(
        """SELECT s.name,s.number,s.version,s.priority,d.original_name FROM standards s
           JOIN documents d ON d.id=s.document_id WHERE s.project_id=? ORDER BY s.priority""", (project_id,))
    issue_rows = [{
        "序号": index, "项目": project["name"], "供应商": project["supplier"], "产品": project["product_name"],
        "问题等级": row["severity"], "问题类别": row["category"], "文件名称": row["source_file"],
        "页码": row["source_page"], "检查项目": row["item"], "实际内容": row["actual"],
        "要求内容": row["requirement"], "对应标准": row["standard_file"], "标准条款": row["standard_clause"],
        "问题描述": row["description"], "整改建议": row["suggestion"], "AI置信度": row["confidence"],
        "人工状态": row["status"], "备注": "",
    } for index, row in enumerate(findings, start=1)]
    counts = {level: sum(row["severity"] == level for row in findings) for level in ("Critical", "Major", "Minor", "Warning", "Review")}
    summary = [{"指标": "文件数量", "数量": len(documents)}, {"指标": "检查数量", "数量": len(findings)},
               {"指标": "合格数量", "数量": "仅统计问题，暂不推算"}, *({"指标": key, "数量": value} for key, value in counts.items())]
    basis = [{"文件名称": row["original_name"], "标准名称": row["name"], "标准编号": row["number"],
              "版本": row["version"], "优先级": row["priority"]} for row in standards]
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(issue_rows, columns=["序号", "项目", "供应商", "产品", "问题等级", "问题类别", "文件名称", "页码",
            "检查项目", "实际内容", "要求内容", "对应标准", "标准条款", "问题描述", "整改建议", "AI置信度", "人工状态", "备注"]).to_excel(writer, sheet_name="问题清单", index=False)
        pd.DataFrame(summary).to_excel(writer, sheet_name="审核汇总", index=False)
        pd.DataFrame(basis, columns=["文件名称", "标准名称", "标准编号", "版本", "优先级"]).to_excel(writer, sheet_name="审核依据", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center")
            for column in sheet.columns:
                values = [len(str(cell.value or "")) for cell in column]
                sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(values + [8]) + 2, 45)
                for cell in column: cell.alignment = Alignment(vertical="top", wrap_text=True)
    return output.getvalue()


def export_batch(db: ReviewDatabase, batch_id: str) -> bytes:
    batch = db.one("SELECT b.*,t.name template_name FROM review_batches b LEFT JOIN audit_templates t ON t.id=b.template_id WHERE b.id=?", (batch_id,))
    if not batch:
        raise ValueError("审核批次不存在")
    findings = db.query("SELECT * FROM findings WHERE batch_id=? ORDER BY id", (batch_id,))
    documents = db.query(
        """SELECT d.*,bd.role,bd.priority FROM batch_documents bd JOIN documents d ON d.id=bd.document_id
           WHERE bd.batch_id=? ORDER BY bd.priority,d.created_at""", (batch_id,)
    )
    rule_evaluations = db.query(
        """SELECT task_index,task_name,status,conclusion,source_file,source_page,evidence,evidence_type,
                  actual,requirement,logic,suggestion,confidence,error,metadata,started_at,completed_at
           FROM rule_evaluations WHERE batch_id=? ORDER BY task_index""", (batch_id,)
    )
    def downgrade_reason(finding: dict[str, object]) -> str:
        try:
            reasons = json.loads(str(finding.get("metadata") or "{}")).get("downgrade_reasons", [])
        except (ValueError, TypeError):
            reasons = []
        return "；".join(str(reason) for reason in reasons)

    rows = [{
        "序号": index, "审核批次": batch["name"], "审核模板": batch.get("template_name") or "",
        "供应商": batch.get("supplier_name") or "",
        "问题等级": finding["severity"], "问题类别": finding["category"], "文件名称": finding["source_file"],
        "页码": finding["source_page"], "检查项目": finding["item"], "实际内容": finding["actual"],
        "要求内容": finding["requirement"], "对应标准": finding["standard_file"],
        "标准页码": finding["standard_page"], "标准条款": finding["standard_clause"],
        "判断逻辑": finding["logic"], "问题描述": finding["description"], "整改建议": finding["suggestion"],
        "AI置信度": finding["confidence"], "降级原因": downgrade_reason(finding),
        "人工状态": finding["status"], "备注": "",
    } for index, finding in enumerate(findings, start=1)]
    levels = ("Critical", "Major", "Minor", "Warning", "Review")
    summary = [{"指标": "审核批次", "内容": batch["name"]}, {"指标": "供应商", "内容": batch.get("supplier_name") or "未识别"},
               {"指标": "审核模板", "内容": batch.get("template_name") or "-"},
               {"指标": "文件数量", "内容": len(documents)},
               {"指标": "问题数量", "内容": len(findings)},
               *({"指标": level, "内容": sum(item["severity"] == level for item in findings)} for level in levels)]
    basis = [{"文件名称": item["original_name"], "类别": item["document_kind"], "角色": item["role"],
              "优先级": item["priority"], "页数": item["page_count"], "解析状态": item["parse_status"],
              "依据检索": "本地关键词"} for item in documents if item["role"] != "supplier"]
    rule_rows = [{
        "序号": item["task_index"], "审核任务": item["task_name"], "结论": item["status"],
        "结论说明": item["conclusion"], "证据文件": item["source_file"], "页码": item["source_page"],
        "证据": item["evidence"], "证据类型": item["evidence_type"], "实际值": item["actual"],
        "要求值": item["requirement"], "判断逻辑": item["logic"], "整改建议": item["suggestion"],
        "置信度": item["confidence"], "降级原因": downgrade_reason(item), "调用错误": item["error"], "开始时间": item["started_at"],
        "完成时间": item["completed_at"],
    } for item in rule_evaluations]
    output = BytesIO()
    columns = ["序号", "审核批次", "审核模板", "供应商", "问题等级", "问题类别", "文件名称", "页码", "检查项目", "实际内容",
               "要求内容", "对应标准", "标准页码", "标准条款", "判断逻辑", "问题描述", "整改建议", "AI置信度", "降级原因", "人工状态", "备注"]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(rows, columns=columns).to_excel(writer, sheet_name="问题清单", index=False)
        pd.DataFrame(summary).to_excel(writer, sheet_name="审核汇总", index=False)
        pd.DataFrame(rule_rows, columns=["序号", "审核任务", "结论", "结论说明", "证据文件", "页码", "证据", "证据类型",
            "实际值", "要求值", "判断逻辑", "整改建议", "置信度", "降级原因", "调用错误", "开始时间", "完成时间"]).to_excel(
                writer, sheet_name="逐条规则", index=False)
        pd.DataFrame(basis, columns=["文件名称", "类别", "角色", "优先级", "页数", "解析状态", "依据检索"]).to_excel(writer, sheet_name="审核依据", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="111827")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in sheet.columns:
                width = min(max([len(str(cell.value or "")) for cell in column] + [8]) + 2, 48)
                sheet.column_dimensions[get_column_letter(column[0].column)].width = width
                for cell in column: cell.alignment = Alignment(vertical="top", wrap_text=True)
    return output.getvalue()
