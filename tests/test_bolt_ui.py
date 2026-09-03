"""UI regression tests use an isolated database, never the user's review queue."""
import json
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

import pytest
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from app.auditing.bolt_template import BOLT_TEMPLATE_NAME, SINGLE_NOTICE
from app.auditing.v2_service import ReviewService
from app.database import ReviewDatabase
from app.database.v2 import utcnow
from app.exporters import export_batch
from app.integrations import ConfigStore


@pytest.fixture
def context(tmp_path):
    db = ReviewDatabase(tmp_path/'review.db')
    store = ConfigStore(db, tmp_path/'key')
    service = ReviewService(db, store, tmp_path/'uploads', tmp_path/'basis', tmp_path/'vectors')
    return SimpleNamespace(db=db, config_store=store, service=service)


@pytest.mark.parametrize('page', ['start_review_page', 'templates_page', 'review_records_page'])
def test_pages_render_without_errors(context, page):
    with patch('app.ui.pages.get_context', return_value=context):
        app = AppTest.from_string(f'from app.ui.pages import {page}\n{page}()').run()
        assert not app.exception


def test_single_document_scope_and_serial_execution_are_visible(context):
    with patch('app.ui.pages.get_context', return_value=context):
        app = AppTest.from_string('from app.ui.pages import start_review_page\nstart_review_page()').run()
        app.selectbox[0].select(BOLT_TEMPLATE_NAME).run()
        app.selectbox[1].select('单文件预检查').run()
        assert not app.exception
        assert any(SINGLE_NOTICE in item.value for item in app.warning)
        assert app.toggle(key='review_parallel_enabled').disabled


def test_template_save_retains_structured_binding_and_disabled_rules(context):
    with patch('app.ui.pages.get_context', return_value=context):
        app = AppTest.from_string('from app.ui.pages import templates_page\ntemplates_page()').run()
        app.selectbox[0].select(BOLT_TEMPLATE_NAME).run()
        template = context.db.one('SELECT * FROM audit_templates WHERE name=?', (BOLT_TEMPLATE_NAME,))
        before = json.loads(template['required_items'])
        app.text_area(key=f"template_instructions_{template['id']}").set_value('User-owned instructions').run()
        app.button(key=f"template_save_{template['id']}").click().run()
        assert not app.exception
        saved = context.db.one('SELECT * FROM audit_templates WHERE id=?', (template['id'],))
        assert json.loads(saved['required_items']) == before
        assert saved['review_instructions'] == 'User-owned instructions'


def test_excel_includes_visual_evidence_when_present(context):
    db = context.db
    batch = db.create_batch(context.service.ensure_bolt_template(), audit_scope='single_document')
    did = db.add_document(library='supplier',kind='supplier',original_name='COC.pdf',stored_path='/unused.pdf',sha256='fixture')
    db.attach_document(batch,did,'supplier',4)
    db.execute('''INSERT INTO visual_evidence(batch_id,document_id,page,bbox,kind,state,method,confidence,
        details,model_fingerprint,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)''',
        (batch,did,1,'[10,20,30,40]','signature','present','local_vision',.95,'{}','synthetic',utcnow()))
    book = load_workbook(BytesIO(export_batch(db,batch)), read_only=True)
    assert '签章证据' in book.sheetnames
    assert book['签章证据']['D2'].value == 'present'
    assert any(row[0] == '审核范围' and row[1] == '单文件预检查' for row in book['审核汇总'].values)


def test_learning_priority_sorts_within_severity_only():
    from app.ui.pages import _finding_review_rank
    rows = [
        {'id':1,'severity':'Review','metadata':'{}'},
        {'id':2,'severity':'Review','metadata':'{"review_priority":"high"}'},
        {'id':3,'severity':'Major','metadata':'malformed'},
    ]
    assert [row['id'] for row in sorted(rows,key=_finding_review_rank)] == [3,2,1]
