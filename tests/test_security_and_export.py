from io import BytesIO
from pathlib import Path

import pytest

from app.database import Database
from app.exporters import export_findings
from app.models import Finding
from app.utils import safe_filename, save_upload


def test_safe_upload_blocks_path_traversal_and_extensions(tmp_path: Path) -> None:
    assert safe_filename("../../报告.pdf") == "报告.pdf"
    with pytest.raises(ValueError): safe_filename("payload.exe")
    saved = save_upload(BytesIO(b"%PDF fake"), "../ok.pdf", tmp_path, 100)
    assert saved.parent == tmp_path


def test_export_contains_three_sheets(tmp_path: Path) -> None:
    from openpyxl import load_workbook
    db = Database(tmp_path / "app.db")
    project = db.create_project({"name": "P1", "supplier": "S1"})
    db.add_finding(project, Finding("数值不符合", "Major", "抗拉强度", "低于要求", "455 MPa", ">=470 MPa"))
    output = tmp_path / "result.xlsx"; output.write_bytes(export_findings(db, project))
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["问题清单", "审核汇总", "审核依据"]
    workbook.close()

